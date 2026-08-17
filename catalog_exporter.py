from __future__ import annotations

import hashlib
import io
import math
import os
import re
import shutil
import subprocess
import tempfile
import time
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from functools import lru_cache
from pathlib import Path
from typing import Callable, Literal
from urllib.parse import urlparse
from urllib.request import Request, urlopen

from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont
from pypdf import PdfReader, PdfWriter, Transformation
from reportlab.lib.colors import HexColor
from reportlab.lib.utils import ImageReader
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfgen import canvas


APP_DIR = Path(__file__).resolve().parent
DEFAULT_POR_BADGE_IMAGE_FILE = APP_DIR / "templates/por img.png"
DEFAULT_NEW_BADGE_IMAGE_FILE = APP_DIR / "templates/new img.png"
DEFAULT_P6_POR_PANEL_IMAGE_FILE = APP_DIR / "templates/por img p6.png"
CACHE_DIR = Path(tempfile.gettempdir()) / "excel_catalog_exporter_cache"
IMAGE_CACHE_DIR = CACHE_DIR / "images"
PDF_CACHE_DIR = CACHE_DIR / "pdf"
USER_AGENT = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
TITLE_COLOR = "#E31B23"
REFERENCE_PAGE_SIZE = (2193, 3322)
REFERENCE_COLUMNS = [(46, 733), (753, 1440), (1458, 2145)]
REFERENCE_PAGE1_ROWS = [(809, 1392), (1416, 1999), (2014, 2597), (2621, 3203)]
REFERENCE_PAGE2_ROWS = [(200, 783), (809, 1392), (1416, 1999), (2014, 2597), (2621, 3203)]
REFERENCE_LAST_ROWS = [(200, 783), (809, 1392), (1416, 1999), (2014, 2597)]
P6_FIRST_PAGE_ROWS = [(1416, 1999), (2014, 2597), (2621, 3203)]
P6_MIDDLE_PAGE_ROWS = [(809, 1392), (1416, 1999), (2014, 2597), (2621, 3203)]
P6_LAST_PAGE_ROWS = [(809, 1392), (1416, 1999), (2014, 2597)]
P6_REFERENCE_CARD_SIZE = (437, 372)
P6_VERTICAL_IMAGE_BOX = (24, 76, 206, 322)
P6_VERTICAL_TITLE_BOX = (226, 164, 420, 318)
P6_VERTICAL_PACK_BOX = (234, 286, 382, 334)
P6_HORIZONTAL_TITLE_BOX = (44, 38, 282, 138)
P6_HORIZONTAL_PACK_BOX = (44, 126, 282, 180)
P6_HORIZONTAL_IMAGE_BOX = (86, 198, 356, 322)
P6_LABEL_BOX = (292, 8, 427, 168)
P6_LABEL_REFERENCE_SIZE = (135, 160)
P6_NEW_BADGE_BOX = (8, 8, 128, 90)
P6_LABEL_COVER_BOX = (4, 28, 131, 84)
P6_LABEL_PRICE_BOX = (8, 2, 126, 44)
P6_LABEL_SELL_BOX = (8, 42, 126, 76)
P6_LABEL_VALUE_BG_BOX = (4, 90, 88, 126)
P6_LABEL_VALUE_BOX = (7, 93, 85, 123)
P6_LABEL_POR_WORD_BOX = (96, 92, 133, 124)
P6_CARD_INSET_RATIO = 0.018
REFERENCE_LAST_PAGE_FOOTER_MARGIN = 24
REFERENCE_SLOT = (
    REFERENCE_COLUMNS[0][0],
    REFERENCE_PAGE2_ROWS[0][0],
    REFERENCE_COLUMNS[0][1],
    REFERENCE_PAGE2_ROWS[0][1],
)
REFERENCE_PRODUCT_IMAGE_BOX = (54, 26, 518, 408)
REFERENCE_TITLE_BOX = (46, 428, 642, 548)
REFERENCE_BADGE_SIZE = (152, 96)
REFERENCE_BADGE_MARGIN = (19, 18)
REFERENCE_NEW_BADGE_WIDTH = 178
REFERENCE_NEW_BADGE_MARGIN_TOP = 14
REFERENCE_TITLE_HEIGHT = 88
REFERENCE_TITLE_MARGIN_X = 30
REFERENCE_TITLE_BOTTOM_MARGIN = 2
REFERENCE_CONTENT_SIDE_MARGIN = 3
REFERENCE_CONTENT_TOP_MARGIN = 5
REFERENCE_IMAGE_GAP_TO_BADGES = 3
REFERENCE_IMAGE_GAP_TO_TITLE = 1
REFERENCE_POR_TEXT_BOX = (12, 14, 500, 292)
NORMAL_IMAGE_SCALE_PERCENT = 100
VERTICAL_IMAGE_SCALE_PERCENT = 128
VERTICAL_IMAGE_ASPECT_RATIO = 1.45
WHITE_TRIM_THRESHOLD = 245
HIGH_QUALITY_OVERLAY_SCALE = 2
TEMPLATE_ANALYSIS_RENDER_DPI = 72
NORMAL_EXPORT_RENDER_DPI = 150
HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "image_src": ("image src",),
    "title": ("title",),
    "por": (
        "metafield: custom.tier_c_profit_on_return [number_decimal]",
        "metafield: custom.promotion_profit_on_return [number_decimal]",
    ),
    "created_at": ("created at",),
    "main_price": ("metafield: custom.promotion_case_price [number_decimal]",),
    "sale_price": ("metafield: custom.retail_sale_price_rsp [number_decimal]",),
    "case_size": ("metafield: custom.case_size [number_integer]", "case", "case size"),
    "pack_size": ("pack_size", "pack size"),
    "image_position": ("image position",),
}
REQUIRED_HEADERS = {
    "image_src": "Image Src",
    "title": "Title",
    "por": "Profit on Return",
}
ExportQuality = Literal["normal", "high"]
TemplateFormName = Literal["auto", "por-title", "promotion"]
TemplateVariant = Literal["default", "p6"]

TEMPLATE_FORM_VARIANTS: dict[str, TemplateVariant] = {
    "por-title": "default",
    "promotion": "p6",
}

TEMPLATE_FORM_NOTES: dict[str, dict[str, object]] = {
    "por-title": {
        "description": "W15/Week 18 style cards using product image, title, POR badge, and NEW badge only.",
        "product_fields": ("Image Src", "Title", "Created At/NEW marker", "POR"),
        "image_alignment": "Trim whitespace, center horizontally, bottom-align inside a large card image area.",
        "por_badge": {
            "position": "top-right",
            "reference_size": REFERENCE_BADGE_SIZE,
            "reference_margin": REFERENCE_BADGE_MARGIN,
        },
        "new_badge": {
            "position": "top-left",
            "reference_width": REFERENCE_NEW_BADGE_WIDTH,
            "reference_margin_top": REFERENCE_NEW_BADGE_MARGIN_TOP,
        },
        "title": {
            "position": "bottom-centered",
            "reference_height": REFERENCE_TITLE_HEIGHT,
            "reference_margin_x": REFERENCE_TITLE_MARGIN_X,
        },
    },
    "promotion": {
        "description": "Promotion-style cards (P7 BBQ and similar) with price, sale price, POR panel, title, pack size, and horizontal/vertical image-position handling. Uses the same card layout as p6.",
    },
}

POR_BADGE_IMAGE_FILE = DEFAULT_POR_BADGE_IMAGE_FILE
NEW_BADGE_IMAGE_FILE = DEFAULT_NEW_BADGE_IMAGE_FILE
POR_BADGE_P6_IMAGE_FILE = DEFAULT_P6_POR_PANEL_IMAGE_FILE


@dataclass
class Product:
    excel_row: int
    image_source: str | None
    title: str
    por_text: str
    is_new: bool = False
    main_price_text: str = ""
    sale_price_text: str = ""
    pack_size: str = ""
    image_position: str = ""


@dataclass
class PageClassification:
    kind: str
    first_score: float
    middle_score: float


def configure_asset_paths(
    *,
    por_badge_image_file: Path | None = None,
    new_badge_image_file: Path | None = None,
    p6_por_panel_image_file: Path | None = None,
) -> None:
    global POR_BADGE_IMAGE_FILE, NEW_BADGE_IMAGE_FILE, POR_BADGE_P6_IMAGE_FILE

    POR_BADGE_IMAGE_FILE = (por_badge_image_file or DEFAULT_POR_BADGE_IMAGE_FILE).resolve()
    NEW_BADGE_IMAGE_FILE = (new_badge_image_file or DEFAULT_NEW_BADGE_IMAGE_FILE).resolve()
    POR_BADGE_P6_IMAGE_FILE = (p6_por_panel_image_file or DEFAULT_P6_POR_PANEL_IMAGE_FILE).resolve()

    load_rgba_asset.cache_clear()
    load_por_badge_asset.cache_clear()
    load_new_badge_asset.cache_clear()
    load_p6_por_panel_asset.cache_clear()


def try_font(size: int, bold: bool = False) -> ImageFont.FreeTypeFont | ImageFont.ImageFont:
    candidates = [
        "C:/Windows/Fonts/arialbd.ttf" if bold else "C:/Windows/Fonts/arial.ttf",
        "C:/Windows/Fonts/calibrib.ttf" if bold else "C:/Windows/Fonts/calibri.ttf",
        "/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf"
        if bold
        else "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
    ]
    for candidate in candidates:
        if Path(candidate).exists():
            return ImageFont.truetype(candidate, size=size)
    return ImageFont.load_default()


def sanitize_filename(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9]+", "_", value.strip()).strip("_")
    return cleaned or "catalog"


def normalize_header(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def is_new_marker(value: object) -> bool:
    return normalize_header(value) == "new"


def format_por(value: object) -> str:
    raw = str(value or "").strip()
    if not raw:
        return "-"
    raw = raw.replace("%", "")
    try:
        number = Decimal(raw)
    except InvalidOperation:
        return raw if raw.endswith("%") else f"{raw}%"

    quantized = number.quantize(Decimal("0.1")) if number != number.to_integral() else number.quantize(Decimal("1"))
    normalized = format(quantized.normalize(), "f")
    if "." in normalized:
        normalized = normalized.rstrip("0").rstrip(".")
    return f"{normalized}%"


def format_price(value: object) -> str:
    raw = str(value or "").strip()
    if not raw:
        return ""
    raw = raw.replace("£", "").replace(",", "")
    try:
        number = Decimal(raw)
    except InvalidOperation:
        return raw if raw.startswith("£") else f"£{raw}"

    quantized = number.quantize(Decimal("0.01"))
    normalized = format(quantized, "f").rstrip("0").rstrip(".")
    return f"£{normalized}"


def format_case_pack(case_value: object, pack_value: object) -> str:
    case_text = str(case_value or "").strip()
    pack_text = str(pack_value or "").strip()

    if case_text:
        try:
            case_number = Decimal(case_text)
        except InvalidOperation:
            pass
        else:
            formatted = format(
                case_number.quantize(Decimal("1")) if case_number == case_number.to_integral() else case_number.normalize(),
                "f",
            )
            case_text = formatted.rstrip("0").rstrip(".") if "." in formatted else formatted

    if case_text and pack_text:
        return f"{case_text} x {pack_text}"
    return case_text or pack_text


def get_header_indexes(first_row: tuple[object, ...]) -> dict[str, int]:
    return {
        normalize_header(value): index
        for index, value in enumerate(first_row, start=1)
        if normalize_header(value)
    }


def find_header_index(headers: dict[str, int], canonical_name: str) -> int | None:
    for alias in HEADER_ALIASES.get(canonical_name, ()):
        if alias in headers:
            return headers[alias]
    return None


def find_required_header_indexes(headers: dict[str, int]) -> tuple[dict[str, int], list[str]]:
    found: dict[str, int] = {}
    missing: list[str] = []
    for canonical_name, label in REQUIRED_HEADERS.items():
        index = find_header_index(headers, canonical_name)
        if index is None:
            missing.append(label)
        else:
            found[canonical_name] = index
    return found, missing


def normalize_image_position(value: object) -> str:
    text = normalize_header(value)
    if "horiz" in text:
        return "horizontal"
    if "vert" in text:
        return "vertical"
    return ""


def cell_to_text(value: object) -> str:
    return "" if value in (None, "") else str(value).strip()


def resolve_template_variant(
    excel_path: Path,
    template_pdf: Path,
    template_form: TemplateFormName = "auto",
) -> TemplateVariant:
    if template_form != "auto":
        return TEMPLATE_FORM_VARIANTS[template_form]

    try:
        workbook = load_workbook(excel_path, data_only=True, read_only=True)
        sheet = workbook.active
        first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
        workbook.close()
        headers = get_header_indexes(first_row)
        promotion_keys = ("main_price", "sale_price")
        if any(find_header_index(headers, key) is not None for key in promotion_keys):
            return "p6"
    except Exception:
        pass

    return "default"


def load_products(excel_path: Path, template_variant: TemplateVariant = "default") -> list[Product]:
    workbook = load_workbook(excel_path, data_only=True, read_only=True)
    sheet = workbook.active
    first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    headers = get_header_indexes(first_row)
    required_headers, missing = find_required_header_indexes(headers)
    if missing:
        workbook.close()
        raise ValueError(f"Missing required Excel columns: {', '.join(missing)}")

    image_col = required_headers["image_src"]
    title_col = required_headers["title"]
    por_col = required_headers["por"]
    created_at_col = find_header_index(headers, "created_at")
    main_price_col = find_header_index(headers, "main_price")
    sale_price_col = find_header_index(headers, "sale_price")
    case_size_col = find_header_index(headers, "case_size")
    pack_size_col = find_header_index(headers, "pack_size")
    image_position_col = find_header_index(headers, "image_position")

    products: list[Product] = []
    for row_index, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        title = cell_to_text(values[title_col - 1] if title_col - 1 < len(values) else None)
        image_source = cell_to_text(values[image_col - 1] if image_col - 1 < len(values) else None)
        por_value = values[por_col - 1] if por_col - 1 < len(values) else None
        created_at_value = values[created_at_col - 1] if created_at_col and created_at_col - 1 < len(values) else None

        if not title and not image_source and por_value in (None, ""):
            continue
        if image_source == "#VALUE!":
            image_source = ""

        products.append(
            Product(
                excel_row=row_index,
                image_source=image_source or None,
                title=title,
                por_text=format_por(por_value),
                is_new=is_new_marker(created_at_value),
                main_price_text=format_price(values[main_price_col - 1] if main_price_col and main_price_col - 1 < len(values) else None),
                sale_price_text=format_price(values[sale_price_col - 1] if sale_price_col and sale_price_col - 1 < len(values) else None),
                pack_size=format_case_pack(
                    values[case_size_col - 1] if case_size_col and case_size_col - 1 < len(values) else None,
                    values[pack_size_col - 1] if pack_size_col and pack_size_col - 1 < len(values) else None,
                ),
                image_position=normalize_image_position(
                    values[image_position_col - 1] if image_position_col and image_position_col - 1 < len(values) else None
                ),
            )
        )

    workbook.close()
    return products


def build_output_path(excel_path: Path, template_pdf: Path, output_path: Path | None = None) -> Path:
    if output_path is not None:
        return output_path
    filename = f"{sanitize_filename(excel_path.stem)}_{sanitize_filename(template_pdf.stem)}_filled.pdf"
    return excel_path.resolve().parent / filename


def ensure_cache_dir(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)


def clear_directory(path: Path) -> None:
    ensure_cache_dir(path)
    for child in path.iterdir():
        if child.is_file():
            child.unlink()
        elif child.is_dir():
            shutil.rmtree(child, ignore_errors=True)


@lru_cache(maxsize=None)
def find_executable(name: str) -> str:
    executable = shutil.which(name)
    if executable:
        return executable
    raise FileNotFoundError(f"Could not find '{name}'. Install Poppler and add it to PATH.")


def subprocess_window_kwargs() -> dict[str, object]:
    if os.name != "nt":
        return {}
    startupinfo = subprocess.STARTUPINFO()
    startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW
    return {
        "creationflags": getattr(subprocess, "CREATE_NO_WINDOW", 0),
        "startupinfo": startupinfo,
    }


def run_hidden_process(
    args: list[str],
    *,
    check: bool = True,
    capture_output: bool = False,
    text: bool = False,
) -> subprocess.CompletedProcess[str] | subprocess.CompletedProcess[bytes]:
    return subprocess.run(
        args,
        check=check,
        capture_output=capture_output,
        text=text,
        **subprocess_window_kwargs(),
    )


def get_pdf_page_count(pdf_path: Path) -> int:
    with pdf_path.open("rb") as handle:
        return len(PdfReader(handle).pages)


def render_pdf_page(pdf_path: Path, page_number: int, dpi: int = NORMAL_EXPORT_RENDER_DPI) -> Image.Image:
    pdftoppm = find_executable("pdftoppm")
    preview_dir = PDF_CACHE_DIR / f"page_{page_number}_{int(time.time() * 1000)}"
    preview_dir.mkdir(parents=True, exist_ok=True)
    output_root = preview_dir / f"page_{page_number}"
    try:
        run_hidden_process(
            [
                pdftoppm,
                "-r",
                str(dpi),
                "-png",
                "-f",
                str(page_number),
                "-l",
                str(page_number),
                "-singlefile",
                str(pdf_path),
                str(output_root),
            ],
            capture_output=True,
        )
        image_path = output_root.with_suffix(".png")
        with Image.open(image_path) as page_image:
            return page_image.convert("RGB").copy()
    finally:
        shutil.rmtree(preview_dir, ignore_errors=True)


def save_page_pdf(page: Image.Image, output_path: Path) -> None:
    rgb_page = page if page.mode == "RGB" else page.convert("RGB")
    # Optimised JPEG-in-PDF: 4:2:0 chroma subsampling drops colour resolution
    # (invisible on flat catalog art / product photos) while keeping luminance
    # detail for text and edges. Combined with a slightly lower quality this
    # roughly halves output size versus the default quality=70 / no subsampling.
    rgb_page.save(
        output_path,
        "PDF",
        resolution=float(NORMAL_EXPORT_RENDER_DPI),
        quality=62,
        optimize=True,
        subsampling=2,
    )


def merge_pdf_files(input_paths: list[Path], output_path: Path) -> None:
    if not input_paths:
        raise RuntimeError("No rendered PDF pages were created.")
    if len(input_paths) == 1:
        shutil.copyfile(input_paths[0], output_path)
        return
    try:
        pdfunite = find_executable("pdfunite")
        run_hidden_process([pdfunite, *[str(path) for path in input_paths], str(output_path)], capture_output=True)
    except (FileNotFoundError, subprocess.CalledProcessError):
        writer = PdfWriter()
        for path in input_paths:
            reader = PdfReader(str(path))
            for page in reader.pages:
                writer.add_page(page)
        with output_path.open("wb") as handle:
            writer.write(handle)


def scale_box(box: tuple[int, int, int, int], page_size: tuple[int, int]) -> tuple[int, int, int, int]:
    ref_width, ref_height = REFERENCE_PAGE_SIZE
    page_width, page_height = page_size
    sx = page_width / ref_width
    sy = page_height / ref_height
    return (
        round(box[0] * sx),
        round(box[1] * sy),
        round(box[2] * sx),
        round(box[3] * sy),
    )


def build_slots(reference_rows: list[tuple[int, int]], page_size: tuple[int, int]) -> list[tuple[int, int, int, int]]:
    return [
        scale_box((left, top, right, bottom), page_size)
        for top, bottom in reference_rows
        for left, right in REFERENCE_COLUMNS
    ]


def get_first_page_slots(page_size: tuple[int, int]) -> list[tuple[int, int, int, int]]:
    return build_slots(REFERENCE_PAGE1_ROWS, page_size)


def get_middle_page_slots(page_size: tuple[int, int]) -> list[tuple[int, int, int, int]]:
    return build_slots(REFERENCE_PAGE2_ROWS, page_size)


def get_last_page_slots(page_size: tuple[int, int]) -> list[tuple[int, int, int, int]]:
    return build_slots(REFERENCE_LAST_ROWS, page_size)


def get_template_slots(template_variant: TemplateVariant, role: str, page_size: tuple[int, int]) -> list[tuple[int, int, int, int]]:
    if template_variant == "p6":
        if role == "first":
            return build_slots(P6_FIRST_PAGE_ROWS, page_size)
        if role == "middle":
            return build_slots(P6_MIDDLE_PAGE_ROWS, page_size)
        return build_slots(P6_LAST_PAGE_ROWS, page_size)
    if role == "first":
        return get_first_page_slots(page_size)
    if role == "middle":
        return get_middle_page_slots(page_size)
    return get_last_page_slots(page_size)


def detect_footer_top_reference(page: Image.Image) -> int | None:
    width, height = page.size
    scan_start = max(0, round(height * 0.65))
    step_x = max(1, width // 240)
    dense_rows: list[int] = []

    for y in range(scan_start, height, max(1, height // 500)):
        total = 0
        non_white = 0
        for x in range(0, width, step_x):
            total += 1
            r, g, b = page.getpixel((x, y))[:3]
            if min(r, g, b) < 245 or max(r, g, b) - min(r, g, b) > 12:
                non_white += 1
        if total and (non_white / total) >= 0.45:
            dense_rows.append(y)

    if not dense_rows:
        return None
    return round((min(dense_rows) / height) * REFERENCE_PAGE_SIZE[1])


def get_footer_safe_last_page_slots(page_size: tuple[int, int], footer_top_reference: int | None) -> list[tuple[int, int, int, int]]:
    if footer_top_reference is None:
        return get_middle_page_slots(page_size)

    reference_top = REFERENCE_PAGE2_ROWS[0][0]
    reference_bottom = REFERENCE_PAGE2_ROWS[-1][1]
    available_bottom = footer_top_reference - REFERENCE_LAST_PAGE_FOOTER_MARGIN
    if available_bottom <= reference_top:
        return get_middle_page_slots(page_size)

    scale = (available_bottom - reference_top) / max(1, reference_bottom - reference_top)
    adjusted_rows = [
        (
            round(reference_top + ((row_top - reference_top) * scale)),
            round(reference_top + ((row_bottom - reference_top) * scale)),
        )
        for row_top, row_bottom in REFERENCE_PAGE2_ROWS
    ]
    return build_slots(adjusted_rows, page_size)


def get_scaled_element_box(box: tuple[int, int, int, int], slot: tuple[int, int, int, int]) -> tuple[int, int, int, int]:
    slot_width = slot[2] - slot[0]
    slot_height = slot[3] - slot[1]
    reference_width = REFERENCE_SLOT[2] - REFERENCE_SLOT[0]
    reference_height = REFERENCE_SLOT[3] - REFERENCE_SLOT[1]
    sx = slot_width / reference_width
    sy = slot_height / reference_height
    return (
        slot[0] + round(box[0] * sx),
        slot[1] + round(box[1] * sy),
        slot[0] + round(box[2] * sx),
        slot[1] + round(box[3] * sy),
    )


def get_scaled_badge_size(slot: tuple[int, int, int, int]) -> tuple[int, int]:
    slot_width = slot[2] - slot[0]
    slot_height = slot[3] - slot[1]
    reference_width = REFERENCE_SLOT[2] - REFERENCE_SLOT[0]
    reference_height = REFERENCE_SLOT[3] - REFERENCE_SLOT[1]
    return (
        max(60, round(REFERENCE_BADGE_SIZE[0] * (slot_width / reference_width))),
        max(40, round(REFERENCE_BADGE_SIZE[1] * (slot_height / reference_height))),
    )


def get_scaled_badge_margin(slot: tuple[int, int, int, int]) -> tuple[int, int]:
    slot_width = slot[2] - slot[0]
    slot_height = slot[3] - slot[1]
    reference_width = REFERENCE_SLOT[2] - REFERENCE_SLOT[0]
    reference_height = REFERENCE_SLOT[3] - REFERENCE_SLOT[1]
    return (
        max(8, round(REFERENCE_BADGE_MARGIN[0] * (slot_width / reference_width))),
        max(8, round(REFERENCE_BADGE_MARGIN[1] * (slot_height / reference_height))),
    )


def get_slot_scale(slot: tuple[int, int, int, int]) -> tuple[float, float]:
    slot_width = slot[2] - slot[0]
    slot_height = slot[3] - slot[1]
    reference_width = REFERENCE_SLOT[2] - REFERENCE_SLOT[0]
    reference_height = REFERENCE_SLOT[3] - REFERENCE_SLOT[1]
    return slot_width / reference_width, slot_height / reference_height


def scale_length(reference_value: int, slot: tuple[int, int, int, int], axis: str, minimum: int = 0) -> int:
    sx, sy = get_slot_scale(slot)
    factor = sx if axis == "x" else sy
    return max(minimum, round(reference_value * factor))


def scale_box_between_sizes(
    box: tuple[int, int, int, int],
    reference_size: tuple[int, int],
    target_size: tuple[int, int],
) -> tuple[int, int, int, int]:
    sx = target_size[0] / reference_size[0]
    sy = target_size[1] / reference_size[1]
    return (
        round(box[0] * sx),
        round(box[1] * sy),
        round(box[2] * sx),
        round(box[3] * sy),
    )


def inset_box(box: tuple[int, int, int, int], inset_x: int, inset_y: int) -> tuple[int, int, int, int]:
    left, top, right, bottom = box
    safe_inset_x = min(inset_x, max(0, (right - left - 2) // 2))
    safe_inset_y = min(inset_y, max(0, (bottom - top - 2) // 2))
    return left + safe_inset_x, top + safe_inset_y, right - safe_inset_x, bottom - safe_inset_y


def expand_box(
    box: tuple[int, int, int, int],
    factor: float,
    bounds: tuple[int, int, int, int],
) -> tuple[int, int, int, int]:
    left, top, right, bottom = box
    width = right - left
    height = bottom - top
    extra_x = round((width * factor - width) / 2)
    extra_y = round((height * factor - height) / 2)
    return (
        max(bounds[0], left - extra_x),
        max(bounds[1], top - extra_y),
        min(bounds[2], right + extra_x),
        min(bounds[3], bottom + extra_y),
    )


def get_p6_card_box(slot: tuple[int, int, int, int]) -> tuple[int, int, int, int]:
    slot_width = max(1, slot[2] - slot[0])
    slot_height = max(1, slot[3] - slot[1])
    inset_x = max(8, round(slot_width * P6_CARD_INSET_RATIO))
    inset_y = max(8, round(slot_height * P6_CARD_INSET_RATIO))
    return inset_box(slot, inset_x, inset_y)


def sanitize_template_page(page: Image.Image, slots: list[tuple[int, int, int, int]]) -> Image.Image:
    cleaned = page.copy()
    draw = ImageDraw.Draw(cleaned)
    inset_x = max(10, round(cleaned.size[0] * 0.005))
    inset_y = max(10, round(cleaned.size[1] * 0.005))
    for left, top, right, bottom in slots:
        draw.rectangle((left + inset_x, top + inset_y, right - inset_x, bottom - inset_y), fill="#FFFFFF")
    return cleaned


def detect_grid_border_style(
    page: Image.Image, slots: list[tuple[int, int, int, int]]
) -> tuple[tuple[int, int, int], int] | None:
    """Sample the real box-border color/thickness from slots known to align with the printed artwork."""
    if not slots:
        return None
    left, top, right, _bottom = slots[0]
    mid_x = min(max(0, (left + right) // 2), page.size[0] - 1)
    scan_start = max(0, top - 20)
    scan_end = min(page.size[1] - 1, top + 20)
    colors: list[tuple[int, int, int]] = []
    for y in range(scan_start, scan_end + 1):
        r, g, b = page.getpixel((mid_x, y))[:3]
        if min(r, g, b) < 230 or max(r, g, b) - min(r, g, b) > 25:
            colors.append((r, g, b))
    if not colors:
        return None
    color = colors[len(colors) // 2]
    return color, max(2, len(colors))


def clear_grid_band(
    page: Image.Image,
    slots: list[tuple[int, int, int, int]],
    bottom_override: int | None = None,
) -> Image.Image:
    """Whiten the full band spanned by slots so stale, misaligned border artwork cannot show through."""
    cleaned = page.copy()
    if not slots:
        return cleaned
    draw = ImageDraw.Draw(cleaned)
    top = min(slot[1] for slot in slots)
    bottom = max(slot[3] for slot in slots)
    if bottom_override is not None:
        bottom = max(bottom, bottom_override)
    margin = max(4, round(cleaned.size[1] * 0.003))
    draw.rectangle((0, max(0, top - margin), cleaned.size[0] - 1, min(cleaned.size[1] - 1, bottom + margin)), fill="#FFFFFF")
    return cleaned


def draw_grid_borders(
    page: Image.Image,
    slots: list[tuple[int, int, int, int]],
    color: tuple[int, int, int],
    width: int,
) -> None:
    draw = ImageDraw.Draw(page)
    for slot in slots:
        draw.rectangle(slot, outline=color, width=width)


def draw_band_cleanup_overlay_pdf(
    page_canvas: canvas.Canvas,
    page_size: tuple[int, int],
    slots: list[tuple[int, int, int, int]],
    bottom_override: int | None = None,
) -> None:
    if not slots:
        return
    top = min(slot[1] for slot in slots)
    bottom = max(slot[3] for slot in slots)
    if bottom_override is not None:
        bottom = max(bottom, bottom_override)
    margin = max(4, round(page_size[1] * 0.003))
    band_box = (0, max(0, top - margin), page_size[0], min(page_size[1], bottom + margin))
    x, y, width, height = to_pdf_rect(band_box, page_size[1])
    page_canvas.setFillColor(HexColor("#FFFFFF"))
    page_canvas.setStrokeColor(HexColor("#FFFFFF"))
    page_canvas.rect(x, y, width, height, stroke=0, fill=1)


def draw_grid_borders_pdf(
    page_canvas: canvas.Canvas,
    slots: list[tuple[int, int, int, int]],
    color: tuple[int, int, int],
    width: int,
    page_height: int,
) -> None:
    page_canvas.setStrokeColorRGB(color[0] / 255, color[1] / 255, color[2] / 255)
    page_canvas.setLineWidth(width)
    for slot in slots:
        x, y, box_width, box_height = to_pdf_rect(slot, page_height)
        page_canvas.rect(x, y, box_width, box_height, stroke=1, fill=0)


def sample_has_border(page: Image.Image, x: int, y: int) -> bool:
    width, height = page.size
    for dx in (-1, 0, 1):
        for dy in (-1, 0, 1):
            px = min(max(0, x + dx), width - 1)
            py = min(max(0, y + dy), height - 1)
            r, g, b = page.getpixel((px, py))[:3]
            if min(r, g, b) < 230 or max(r, g, b) - min(r, g, b) > 25:
                return True
    return False


def slot_border_score(page: Image.Image, slot: tuple[int, int, int, int]) -> float:
    left, top, right, bottom = slot
    inset = max(2, round(min(right - left, bottom - top) * 0.01))
    xs = [left + (right - left) // 5, left + (right - left) // 2, right - (right - left) // 5]
    ys = [top + (bottom - top) // 5, top + (bottom - top) // 2, bottom - (bottom - top) // 5]
    samples: list[tuple[int, int]] = []
    samples.extend((x, top + inset) for x in xs)
    samples.extend((x, bottom - inset) for x in xs)
    samples.extend((left + inset, y) for y in ys)
    samples.extend((right - inset, y) for y in ys)
    hits = sum(1 for x, y in samples if sample_has_border(page, x, y))
    return hits / max(1, len(samples))


def layout_score(page: Image.Image, slots: list[tuple[int, int, int, int]]) -> float:
    scores = [slot_border_score(page, slot) for slot in slots]
    if not scores:
        return 0.0
    strong_hits = sum(1 for score in scores if score >= 0.55)
    return strong_hits / len(scores)


def classify_page_layout(page: Image.Image) -> PageClassification:
    first_score = layout_score(page, get_first_page_slots(page.size))
    top_row_slots = get_middle_page_slots(page.size)[:3]
    middle_score = sum(slot_border_score(page, slot) for slot in top_row_slots) / max(1, len(top_row_slots))
    if middle_score >= 0.72 and first_score >= 0.55:
        kind = "toprow"
    elif first_score >= 0.55:
        kind = "first"
    else:
        kind = "invalid"
    return PageClassification(kind=kind, first_score=first_score, middle_score=middle_score)


def resolve_template_page_numbers(
    template_pdf: Path,
    status_callback: Callable[[str], None] | None = None,
) -> tuple[int, int, int]:
    page_count = get_pdf_page_count(template_pdf)
    if status_callback:
        status_callback(f"Reading template structure ({page_count} page{'s' if page_count != 1 else ''})...")

    if page_count == 1:
        page = render_pdf_page(template_pdf, 1, dpi=TEMPLATE_ANALYSIS_RENDER_DPI)
        try:
            classification = classify_page_layout(page)
        finally:
            page.close()
        if classification.kind == "invalid":
            raise ValueError("Invalid PDF template. The PDF must contain the supported 3-column product box grid.")
        return 1, 1, 1

    if page_count == 2:
        first_page = render_pdf_page(template_pdf, 1, dpi=TEMPLATE_ANALYSIS_RENDER_DPI)
        second_page = render_pdf_page(template_pdf, 2, dpi=TEMPLATE_ANALYSIS_RENDER_DPI)
        try:
            first_class = classify_page_layout(first_page)
            second_class = classify_page_layout(second_page)
        finally:
            close_unique_images(first_page, second_page)
        if {first_class.kind, second_class.kind} != {"first", "toprow"}:
            raise ValueError(
                "Invalid 2-page PDF template. It must contain one cover/end-style page and one middle-style 3-column grid page."
            )
        first_index = 1 if first_class.kind == "first" else 2
        middle_index = 1 if first_class.kind == "toprow" else 2
        return first_index, middle_index, first_index

    first_page = render_pdf_page(template_pdf, 1, dpi=TEMPLATE_ANALYSIS_RENDER_DPI)
    middle_page = render_pdf_page(template_pdf, 2, dpi=TEMPLATE_ANALYSIS_RENDER_DPI)
    last_page = render_pdf_page(template_pdf, page_count, dpi=TEMPLATE_ANALYSIS_RENDER_DPI)
    try:
        first_class = classify_page_layout(first_page)
        middle_class = classify_page_layout(middle_page)
        last_class = classify_page_layout(last_page)
    finally:
        close_unique_images(first_page, middle_page, last_page)
    if first_class.kind == "invalid" or middle_class.kind == "invalid" or last_class.kind == "invalid":
        raise ValueError("Invalid PDF template. The PDF must contain the supported 3-column product box grid.")
    if middle_class.kind != "toprow":
        raise ValueError("Invalid PDF template. Page 2 must be a middle-style 3-column grid page.")
    return 1, 2, page_count


def load_template_themes(
    template_pdf: Path,
    status_callback: Callable[[str], None] | None = None,
    page_numbers: tuple[int, int, int] | None = None,
) -> tuple[Image.Image, Image.Image, Image.Image]:
    first_index, middle_index, last_index = page_numbers or resolve_template_page_numbers(
        template_pdf,
        status_callback=status_callback,
    )
    rendered_pages: dict[int, Image.Image] = {}
    for page_number in {first_index, middle_index, last_index}:
        rendered_pages[page_number] = render_pdf_page(template_pdf, page_number, dpi=NORMAL_EXPORT_RENDER_DPI)
    return (
        rendered_pages[first_index],
        rendered_pages[middle_index],
        rendered_pages[last_index],
    )


def cache_path_for_source(source: str) -> Path:
    parsed = urlparse(source)
    suffix = Path(parsed.path).suffix or ".img"
    digest = hashlib.sha1(source.encode("utf-8")).hexdigest()
    return IMAGE_CACHE_DIR / f"{digest}{suffix}"


def download_image(source: str) -> bytes:
    request = Request(source, headers={"User-Agent": USER_AGENT})
    with urlopen(request, timeout=30) as response:
        return response.read()


def load_product_image(excel_path: Path, source: str | None) -> Image.Image | None:
    if not source:
        return None

    try:
        if source.startswith(("http://", "https://")):
            ensure_cache_dir(IMAGE_CACHE_DIR)
            cached_path = cache_path_for_source(source)
            if cached_path.exists():
                with Image.open(cached_path) as cached_image:
                    return cached_image.convert("RGBA")

            data = download_image(source)
            cached_path.write_bytes(data)
            with Image.open(io.BytesIO(data)) as downloaded_image:
                return downloaded_image.convert("RGBA")

        candidate = Path(source)
        if not candidate.is_absolute():
            candidate = (excel_path.parent / candidate).resolve()
        if candidate.exists():
            with Image.open(candidate) as local_image:
                return local_image.convert("RGBA")
    except Exception as error:
        print(f"Warning: could not load image '{source}': {error}")

    return None


@lru_cache(maxsize=4)
def load_rgba_asset(asset_path: str) -> Image.Image | None:
    path = Path(asset_path)
    if not path.exists():
        return None
    with Image.open(path) as asset:
        return asset.convert("RGBA").copy()


@lru_cache(maxsize=1)
def load_por_badge_asset() -> Image.Image | None:
    return load_rgba_asset(str(POR_BADGE_IMAGE_FILE))


@lru_cache(maxsize=1)
def load_new_badge_asset() -> Image.Image | None:
    asset = load_rgba_asset(str(NEW_BADGE_IMAGE_FILE))
    if asset is None:
        return None
    left_half = asset.crop((0, 0, asset.width // 2, asset.height))
    alpha_box = left_half.getchannel("A").getbbox()
    return left_half.crop(alpha_box) if alpha_box else left_half


@lru_cache(maxsize=1)
def load_p6_por_panel_asset() -> Image.Image | None:
    return load_rgba_asset(str(POR_BADGE_P6_IMAGE_FILE))


def create_placeholder(size: tuple[int, int]) -> Image.Image:
    placeholder = Image.new("RGBA", size, (255, 255, 255, 0))
    draw = ImageDraw.Draw(placeholder)
    border = max(3, round(min(size) * 0.01))
    draw.rounded_rectangle((0, 0, size[0] - 1, size[1] - 1), radius=max(16, round(min(size) * 0.07)), outline="#D8D8D8", width=border)
    draw.line((size[0] * 0.2, size[1] * 0.25, size[0] * 0.8, size[1] * 0.75), fill="#D8D8D8", width=border)
    draw.line((size[0] * 0.8, size[1] * 0.25, size[0] * 0.2, size[1] * 0.75), fill="#D8D8D8", width=border)
    message_font = try_font(max(14, round(min(size) * 0.06)), bold=True)
    draw.text((size[0] // 2, size[1] - max(26, round(size[1] * 0.09))), "IMAGE UNAVAILABLE", anchor="mm", fill="#8E8E8E", font=message_font)
    return placeholder


def trim_product_image(image: Image.Image) -> Image.Image:
    rgba = image if image.mode == "RGBA" else image.convert("RGBA")
    alpha_bbox = rgba.getchannel("A").getbbox()
    trimmed = rgba.crop(alpha_bbox) if alpha_bbox else rgba.copy()
    non_white_mask = trimmed.convert("L").point(lambda value: 255 if value < WHITE_TRIM_THRESHOLD else 0)
    content_bbox = non_white_mask.getbbox()
    return trimmed.crop(content_bbox) if content_bbox else trimmed


def fit_text_block(
    draw: ImageDraw.ImageDraw,
    text: str,
    box: tuple[int, int, int, int],
    *,
    bold: bool = True,
    min_size: int = 14,
    max_lines: int = 4,
) -> tuple[ImageFont.ImageFont, list[str], int]:
    display_text = re.sub(r"\s+", " ", text.strip()).upper() if bold else re.sub(r"\s+", " ", text.strip())
    display_text = display_text or "UNTITLED PRODUCT"
    max_width = max(1, box[2] - box[0])
    max_height = max(1, box[3] - box[1])
    start_size = max(min_size + 4, round(max_height * 0.24))
    end_size = max(10, min_size)
    for size in range(start_size, end_size - 1, -1):
        font = try_font(size, bold=bold)
        lines = wrap_text(draw, display_text, font, max_width)
        if len(lines) > max_lines:
            continue
        line_height = draw.textbbox((0, 0), "Ag", font=font)[3] + max(2, size // 7)
        if line_height * len(lines) <= max_height:
            return font, lines, line_height
    fallback = try_font(end_size, bold=bold)
    lines = wrap_text(draw, display_text, fallback, max_width)[:max_lines]
    return fallback, lines, draw.textbbox((0, 0), "Ag", font=fallback)[3] + max(2, end_size // 7)


def fit_text_block_fixed_range(
    draw: ImageDraw.ImageDraw,
    text: str,
    box: tuple[int, int, int, int],
    *,
    bold: bool = True,
    min_size: int = 18,
    max_size: int = 28,
    max_lines: int = 4,
) -> tuple[ImageFont.ImageFont, list[str], int]:
    display_text = re.sub(r"\s+", " ", text.strip()).upper() if bold else re.sub(r"\s+", " ", text.strip())
    display_text = display_text or "UNTITLED PRODUCT"
    max_width = max(1, box[2] - box[0])
    max_height = max(1, box[3] - box[1])
    for size in range(max_size, min_size - 1, -1):
        font = try_font(size, bold=bold)
        lines = wrap_text(draw, display_text, font, max_width)
        if len(lines) > max_lines:
            continue
        line_height = draw.textbbox((0, 0), "Ag", font=font)[3] + max(2, size // 7)
        if line_height * len(lines) <= max_height:
            return font, lines, line_height
    fallback = try_font(min_size, bold=bold)
    lines = wrap_text(draw, display_text, fallback, max_width)[:max_lines]
    return fallback, lines, draw.textbbox((0, 0), "Ag", font=fallback)[3] + max(2, min_size // 7)


def fit_product_image_p6(
    prepared: Image.Image | None,
    box_size: tuple[int, int],
    *,
    vertical_align: str = "bottom",
) -> Image.Image:
    image = prepared.copy() if prepared is not None else create_placeholder(box_size)
    if image.mode != "RGBA":
        image = image.convert("RGBA")
    image.thumbnail(box_size, Image.Resampling.LANCZOS)

    canvas = Image.new("RGBA", box_size, (255, 255, 255, 0))
    offset_x = (box_size[0] - image.width) // 2
    offset_y = (box_size[1] - image.height) // 2 if vertical_align == "center" else box_size[1] - image.height
    canvas.paste(image, (offset_x, offset_y), image)
    return canvas


def build_p6_price_panel(size: tuple[int, int], product: Product) -> Image.Image:
    asset = load_p6_por_panel_asset()
    if asset is None:
        width, height = size
        panel = Image.new("RGBA", size, (0, 0, 0, 0))
        draw = ImageDraw.Draw(panel)
        draw.polygon(
            [(0, 0), (width - 1, 0), (width - 1, height - max(12, round(height * 0.11))), (0, height - 1)],
            fill="#D91E24",
        )
        draw.rectangle((0, round(height * 0.72), width - 1, round(height * 0.80)), fill="#C81A20")
    else:
        panel = asset.resize(size, Image.Resampling.LANCZOS)
        draw = ImageDraw.Draw(panel)

    price_box = scale_box_between_sizes(P6_LABEL_PRICE_BOX, P6_LABEL_REFERENCE_SIZE, size)
    sell_box = scale_box_between_sizes(P6_LABEL_SELL_BOX, P6_LABEL_REFERENCE_SIZE, size)
    value_bg_box = scale_box_between_sizes(P6_LABEL_VALUE_BG_BOX, P6_LABEL_REFERENCE_SIZE, size)
    value_box = scale_box_between_sizes(P6_LABEL_VALUE_BOX, P6_LABEL_REFERENCE_SIZE, size)
    por_word_box = scale_box_between_sizes(P6_LABEL_POR_WORD_BOX, P6_LABEL_REFERENCE_SIZE, size)

    # Replace the source asset's top POR pill with a flat red field like the reference.
    cover_bottom = scale_box_between_sizes(P6_LABEL_COVER_BOX, P6_LABEL_REFERENCE_SIZE, size)[3]
    cover_color = panel.getpixel((max(1, size[0] // 20), max(1, size[1] // 20)))
    draw.rectangle((0, 0, size[0], cover_bottom), fill=cover_color)

    price_text = product.main_price_text or "\xA30"
    sell_text = f"SELL {product.sale_price_text}" if product.sale_price_text else "SELL"
    por_text = product.por_text or "-"

    price_font = fit_single_line_font_custom(draw, price_text, price_box, bold=True, start_factor=1.0, end_factor=0.42)
    sell_font = fit_single_line_font_custom(draw, sell_text, sell_box, bold=True, start_factor=0.82, end_factor=0.36)
    value_font = fit_single_line_font_custom(draw, por_text, value_box, bold=True, start_factor=1.12, end_factor=0.42)
    por_font = fit_single_line_font(draw, "POR", por_word_box, bold=True)

    center_x = (price_box[0] + price_box[2]) // 2
    price_center_y = (price_box[1] + price_box[3]) // 2
    sell_center_y = (sell_box[1] + sell_box[3]) // 2

    pill_height = value_bg_box[3] - value_bg_box[1]
    pill_width = value_bg_box[2] - value_bg_box[0]
    por_center_y = (value_bg_box[1] + value_bg_box[3]) // 2
    por_gap = max(6, round(size[0] * 0.025))
    por_text_bbox = draw.textbbox((0, 0), "POR", font=por_font)
    por_text_width = por_text_bbox[2] - por_text_bbox[0]
    group_width = pill_width + por_gap + por_text_width
    group_left = round(center_x - (group_width / 2))
    pill_left = group_left
    pill_top = round(por_center_y - (pill_height / 2))
    pill_right = pill_left + pill_width
    pill_bottom = pill_top + pill_height
    por_anchor_x = pill_right + por_gap

    draw.rounded_rectangle(
        (pill_left, pill_top, pill_right, pill_bottom),
        radius=max(8, round(pill_height * 0.22)),
        fill="#FFFFFF",
    )

    draw.text(
        (center_x, price_center_y),
        price_text,
        anchor="mm",
        fill="#FFFFFF",
        font=price_font,
    )
    draw.text(
        (center_x, sell_center_y),
        sell_text,
        anchor="mm",
        fill="#FFFFFF",
        font=sell_font,
    )
    draw.text(
        ((pill_left + pill_right) // 2, por_center_y),
        por_text,
        anchor="mm",
        fill="#C11F25",
        font=value_font,
    )
    draw.text(
        (por_anchor_x, por_center_y),
        "POR",
        anchor="lm",
        fill="#FFFFFF",
        font=por_font,
    )
    return panel


def render_p6_card(product: Product, product_image: Image.Image | None, slot: tuple[int, int, int, int]) -> Image.Image:
    slot_width = max(1, slot[2] - slot[0])
    slot_height = max(1, slot[3] - slot[1])
    card = Image.new("RGBA", (slot_width, slot_height), (255, 255, 255, 255))
    draw = ImageDraw.Draw(card)

    image_is_horizontal = product.image_position == "horizontal"

    draw.rectangle((0, 0, slot_width - 1, slot_height - 1), fill="#FFFFFF")

    prepared_image = trim_product_image(product_image) if product_image is not None else None
    inferred_shape = classify_image_shape(prepared_image) if prepared_image is not None else "standard"
    effective_horizontal = image_is_horizontal or (product.image_position == "" and inferred_shape != "vertical")

    if effective_horizontal:
        image_box = scale_box_between_sizes(P6_HORIZONTAL_IMAGE_BOX, P6_REFERENCE_CARD_SIZE, (slot_width, slot_height))
        title_box = scale_box_between_sizes(P6_HORIZONTAL_TITLE_BOX, P6_REFERENCE_CARD_SIZE, (slot_width, slot_height))
        pack_box = scale_box_between_sizes(P6_HORIZONTAL_PACK_BOX, P6_REFERENCE_CARD_SIZE, (slot_width, slot_height))
    else:
        image_box = scale_box_between_sizes(P6_VERTICAL_IMAGE_BOX, P6_REFERENCE_CARD_SIZE, (slot_width, slot_height))
        title_box = scale_box_between_sizes(P6_VERTICAL_TITLE_BOX, P6_REFERENCE_CARD_SIZE, (slot_width, slot_height))
        pack_box = scale_box_between_sizes(P6_VERTICAL_PACK_BOX, P6_REFERENCE_CARD_SIZE, (slot_width, slot_height))

    image_box = expand_box(image_box, 1.20, (0, 0, slot_width, slot_height))
    new_badge_box = scale_box_between_sizes(P6_NEW_BADGE_BOX, P6_REFERENCE_CARD_SIZE, (slot_width, slot_height))
    if product.is_new and not effective_horizontal:
        clear_top = new_badge_box[3] + max(8, round(slot_height * 0.022))
        image_box = (
            image_box[0],
            max(image_box[1], clear_top),
            image_box[2],
            image_box[3],
        )
    if product.is_new and effective_horizontal:
        title_bottom = min(image_box[1] - max(36, round(slot_height * 0.10)), max(title_box[3], round(slot_height * 0.43)))
        title_box = (
            title_box[0],
            max(title_box[1], new_badge_box[3] + max(4, round(slot_height * 0.012))),
            title_box[2],
            title_bottom,
        )
        pack_box = (
            title_box[0],
            title_box[3] + max(6, round(slot_height * 0.016)),
            title_box[2],
            min(image_box[1] - max(4, round(slot_height * 0.012)), title_box[3] + max(34, round(slot_height * 0.11))),
        )

    title_top_padding = max(4, round(slot_height * 0.016))
    title_box = (title_box[0], title_box[1] + title_top_padding, title_box[2], title_box[3])

    fitted_image = fit_product_image_p6(
        prepared_image,
        (max(1, image_box[2] - image_box[0]), max(1, image_box[3] - image_box[1])),
        vertical_align="center" if not effective_horizontal else "bottom",
    )
    try:
        card.paste(fitted_image, (image_box[0], image_box[1]), fitted_image)
    finally:
        fitted_image.close()
        if prepared_image is not None:
            prepared_image.close()

    label_box = scale_box_between_sizes(P6_LABEL_BOX, P6_REFERENCE_CARD_SIZE, (slot_width, slot_height))
    price_panel = build_p6_price_panel((label_box[2] - label_box[0], label_box[3] - label_box[1]), product)
    try:
        card.paste(price_panel, (label_box[0], label_box[1]), price_panel)
    finally:
        price_panel.close()

    if product.is_new:
        new_badge = build_new_badge((new_badge_box[2] - new_badge_box[0], new_badge_box[3] - new_badge_box[1]))
        if new_badge is not None:
            try:
                card.paste(new_badge, (new_badge_box[0], new_badge_box[1]), new_badge)
            finally:
                new_badge.close()

    title_font, title_lines, line_height = fit_text_block_fixed_range(
        draw,
        product.title,
        title_box,
        bold=True,
        min_size=20,
        max_size=32,
        max_lines=4,
    )
    title_total_height = line_height * len(title_lines)
    title_y = title_box[1] + max(0, ((title_box[3] - title_box[1]) - title_total_height) // 2)
    title_anchor_x = (title_box[0] + title_box[2]) // 2
    title_top_y = title_y
    for line in title_lines:
        draw.text((title_anchor_x, title_y), line, anchor="ma", fill="#1F2A37", font=title_font)
        title_y += line_height
    title_bottom_y = title_top_y + title_total_height

    if product.pack_size:
        title_font_size = getattr(title_font, "size", 24)
        pack_target_size = max(24, title_font_size)
        pack_fit_box = (title_box[0], pack_box[1], title_box[2], pack_box[3])
        pack_font = fit_single_line_font_fixed_range(
            draw,
            product.pack_size,
            pack_fit_box,
            bold=True,
            min_size=max(22, pack_target_size - 2),
            max_size=pack_target_size,
        )
        pack_text_box = draw.textbbox((0, 0), product.pack_size, font=pack_font)
        pack_text_height = pack_text_box[3] - pack_text_box[1]
        pack_gap = max(8, round(slot_height * 0.02))
        desired_pack_center_y = round(title_bottom_y + pack_gap + (pack_text_height / 2))
        min_pack_center_y = pack_fit_box[1] + max(1, math.ceil(pack_text_height / 2))
        max_pack_center_y = pack_fit_box[3] - max(1, math.ceil(pack_text_height / 2))
        pack_center_y = min(max_pack_center_y, max(min_pack_center_y, desired_pack_center_y))
        draw.text(
            (title_anchor_x, pack_center_y),
            product.pack_size,
            anchor="mm",
            fill="#111111",
            font=pack_font,
        )

    return card


def classify_image_shape(image: Image.Image) -> str:
    width = max(1, image.width)
    height = max(1, image.height)
    return "vertical" if (height / width) >= VERTICAL_IMAGE_ASPECT_RATIO else "standard"


def fit_product_image(prepared: Image.Image | None, box_size: tuple[int, int], image_shape: str = "standard") -> Image.Image:
    image = prepared.copy() if prepared is not None else create_placeholder(box_size)
    if image.mode != "RGBA":
        image = image.convert("RGBA")
    image.thumbnail(box_size, Image.Resampling.LANCZOS)

    scale_percent = VERTICAL_IMAGE_SCALE_PERCENT if image_shape == "vertical" else NORMAL_IMAGE_SCALE_PERCENT
    if scale_percent != 100:
        scaled = image.resize(
            (
                max(1, round(image.width * (scale_percent / 100))),
                max(1, round(image.height * (scale_percent / 100))),
            ),
            Image.Resampling.LANCZOS,
        )
        if scaled.width <= box_size[0] and scaled.height <= box_size[1]:
            image = scaled

    canvas = Image.new("RGBA", box_size, (255, 255, 255, 0))
    offset_x = (box_size[0] - image.width) // 2
    offset_y = box_size[1] - image.height
    canvas.paste(image, (offset_x, offset_y), image)
    return canvas


def wrap_text(draw: ImageDraw.ImageDraw, text: str, font: ImageFont.ImageFont, max_width: int) -> list[str]:
    words = text.split()
    if not words:
        return [""]
    lines: list[str] = []
    current = words[0]
    for word in words[1:]:
        trial = f"{current} {word}"
        if draw.textbbox((0, 0), trial, font=font)[2] <= max_width:
            current = trial
        else:
            lines.append(current)
            current = word
    lines.append(current)
    return lines


def fit_title(draw: ImageDraw.ImageDraw, text: str, box: tuple[int, int, int, int]) -> tuple[ImageFont.ImageFont, list[str], int]:
    display_text = re.sub(r"\s+", " ", text.strip()).upper() or "UNTITLED PRODUCT"
    max_width = box[2] - box[0]
    max_height = box[3] - box[1]
    start_size = max(20, round(max_height * 0.28))
    end_size = max(14, round(max_height * 0.12))

    for size in range(start_size, end_size - 1, -1):
        font = try_font(size, bold=True)
        lines = wrap_text(draw, display_text, font, max_width)
        if len(lines) > 3:
            continue
        line_height = draw.textbbox((0, 0), "Ag", font=font)[3] + max(4, size // 6)
        if line_height * len(lines) <= max_height:
            return font, lines, line_height

    fallback = try_font(end_size, bold=True)
    lines = wrap_text(draw, display_text, fallback, max_width)
    return fallback, lines[:3], draw.textbbox((0, 0), "Ag", font=fallback)[3] + max(4, end_size // 6)


def fit_single_line_font(
    draw: ImageDraw.ImageDraw,
    text: str,
    box: tuple[int, int, int, int],
    bold: bool = True,
) -> ImageFont.ImageFont:
    max_width = max(1, box[2] - box[0])
    max_height = max(1, box[3] - box[1])
    start_size = max(18, round(max_height * 0.62))
    end_size = max(12, round(max_height * 0.28))
    for size in range(start_size, end_size - 1, -1):
        font = try_font(size, bold=bold)
        bbox = draw.textbbox((0, 0), text, font=font)
        if (bbox[2] - bbox[0]) <= max_width and (bbox[3] - bbox[1]) <= max_height:
            return font
    return try_font(end_size, bold=bold)


def fit_single_line_font_fixed_range(
    draw: ImageDraw.ImageDraw,
    text: str,
    box: tuple[int, int, int, int],
    *,
    bold: bool = True,
    min_size: int = 16,
    max_size: int = 22,
) -> ImageFont.ImageFont:
    max_width = max(1, box[2] - box[0])
    max_height = max(1, box[3] - box[1])
    for size in range(max_size, min_size - 1, -1):
        font = try_font(size, bold=bold)
        bbox = draw.textbbox((0, 0), text, font=font)
        if (bbox[2] - bbox[0]) <= max_width and (bbox[3] - bbox[1]) <= max_height:
            return font
    return try_font(min_size, bold=bold)


def fit_single_line_font_custom(
    draw: ImageDraw.ImageDraw,
    text: str,
    box: tuple[int, int, int, int],
    *,
    bold: bool = True,
    start_factor: float = 1.1,
    end_factor: float = 0.4,
) -> ImageFont.ImageFont:
    max_width = max(1, box[2] - box[0])
    max_height = max(1, box[3] - box[1])
    start_size = max(18, round(max_height * start_factor))
    end_size = max(10, round(max_height * end_factor))
    for size in range(start_size, end_size - 1, -1):
        font = try_font(size, bold=bold)
        bbox = draw.textbbox((0, 0), text, font=font)
        if (bbox[2] - bbox[0]) <= max_width and (bbox[3] - bbox[1]) <= max_height:
            return font
    return try_font(end_size, bold=bold)


def build_por_badge_fallback(por_text: str, size: tuple[int, int]) -> Image.Image:
    width, height = size
    badge = Image.new("RGBA", size, (0, 0, 0, 0))
    draw = ImageDraw.Draw(badge)
    red = "#E31B23"
    deep_red = "#BE171D"
    white = "#FFFFFF"
    number_box_width = width - max(22, round(width * 0.2))
    number_box_height = max(34, round(height * 0.6))

    draw.polygon([(0, 0), (width - 1, 0), (width - 1, height - 1), (0, height - max(10, round(height * 0.18)))], fill=red)
    draw.rectangle((0, 0, number_box_width, number_box_height), fill=white, outline=red, width=max(2, round(width * 0.02)))
    draw.polygon(
        [
            (0, number_box_height),
            (number_box_width, number_box_height),
            (number_box_width, height - max(8, round(height * 0.15))),
            (0, height - 1),
        ],
        fill=red,
    )
    draw.rectangle((number_box_width, 0, width - 1, height - 1), fill=deep_red)

    number_font = try_font(max(16, round(height * 0.34)), bold=True)
    draw.text(
        (number_box_width // 2, number_box_height // 2),
        por_text,
        anchor="mm",
        fill="#B22024",
        font=number_font,
    )

    por_font = try_font(max(9, round(height * 0.16)), bold=True)
    strip_center_x = number_box_width + (width - number_box_width) // 2
    spacing = max(12, round(height * 0.24))
    start_y = max(14, round(height * 0.18))
    for index, letter in enumerate("POR"):
        draw.text((strip_center_x, start_y + (index * spacing)), letter, anchor="mm", fill=white, font=por_font)

    return badge


def build_por_badge(por_text: str, size: tuple[int, int]) -> Image.Image:
    badge_asset = load_por_badge_asset()
    if badge_asset is None:
        return build_por_badge_fallback(por_text, size)

    badge = badge_asset.resize(size, Image.Resampling.LANCZOS)
    draw = ImageDraw.Draw(badge)
    text_box = scale_box_between_sizes(REFERENCE_POR_TEXT_BOX, (600, 428), size)
    font = fit_single_line_font(draw, por_text, text_box, bold=True)
    draw.text(
        ((text_box[0] + text_box[2]) // 2, (text_box[1] + text_box[3]) // 2),
        por_text,
        anchor="mm",
        fill="#B22024",
        font=font,
    )
    return badge


def build_new_badge(size: tuple[int, int]) -> Image.Image | None:
    badge_asset = load_new_badge_asset()
    if badge_asset is None:
        return None
    badge = badge_asset.copy()
    badge.thumbnail(size, Image.Resampling.LANCZOS)
    canvas = Image.new("RGBA", size, (255, 255, 255, 0))
    offset = ((size[0] - badge.width) // 2, (size[1] - badge.height) // 2)
    canvas.paste(badge, offset, badge)
    return canvas


def get_scaled_new_badge_size(slot: tuple[int, int, int, int]) -> tuple[int, int]:
    badge_asset = load_new_badge_asset()
    if badge_asset is None:
        width = scale_length(REFERENCE_NEW_BADGE_WIDTH, slot, "x", minimum=150)
        return width, max(48, round(width * 0.45))
    width = scale_length(REFERENCE_NEW_BADGE_WIDTH, slot, "x", minimum=150)
    height = max(56, round(width * (badge_asset.height / badge_asset.width)))
    return width, height


def get_title_box(slot: tuple[int, int, int, int]) -> tuple[int, int, int, int]:
    margin_x = scale_length(REFERENCE_TITLE_MARGIN_X, slot, "x", minimum=18)
    title_height = scale_length(REFERENCE_TITLE_HEIGHT, slot, "y", minimum=64)
    bottom_margin = scale_length(REFERENCE_TITLE_BOTTOM_MARGIN, slot, "y", minimum=6)
    return (
        slot[0] + margin_x,
        slot[3] - bottom_margin - title_height,
        slot[2] - margin_x,
        slot[3] - bottom_margin,
    )


def get_product_image_box(slot: tuple[int, int, int, int], is_new: bool, image_shape: str = "standard") -> tuple[int, int, int, int]:
    side_margin = scale_length(REFERENCE_CONTENT_SIDE_MARGIN, slot, "x", minimum=18)
    top_margin = scale_length(REFERENCE_CONTENT_TOP_MARGIN, slot, "y", minimum=10)
    gap_to_badges = scale_length(REFERENCE_IMAGE_GAP_TO_BADGES, slot, "y", minimum=8)
    gap_to_title = scale_length(REFERENCE_IMAGE_GAP_TO_TITLE, slot, "y", minimum=8)
    por_size = get_scaled_badge_size(slot)
    por_margin = get_scaled_badge_margin(slot)
    title_box = get_title_box(slot)

    relaxed_top = slot[1] + top_margin
    content_top = max(relaxed_top, slot[1] + por_margin[1] + por_size[1] + gap_to_badges)
    if is_new:
        new_size = get_scaled_new_badge_size(slot)
        new_top = slot[1] + scale_length(REFERENCE_NEW_BADGE_MARGIN_TOP, slot, "y", minimum=8)
        content_top = max(content_top, new_top + new_size[1] + gap_to_badges)

    if image_shape == "vertical" and VERTICAL_IMAGE_SCALE_PERCENT > 100:
        lift_factor = min(1.0, max(0.0, (VERTICAL_IMAGE_SCALE_PERCENT - 100) / 100))
        content_top = round(content_top - ((content_top - relaxed_top) * lift_factor))

    content_bottom = title_box[1] - gap_to_title
    return (
        slot[0] + side_margin,
        content_top,
        slot[2] - side_margin,
        max(content_top + 20, content_bottom),
    )


def draw_product(
    page: Image.Image,
    product: Product,
    product_image: Image.Image | None,
    slot: tuple[int, int, int, int],
    template_variant: TemplateVariant = "default",
) -> None:
    if template_variant == "p6":
        card_box = get_p6_card_box(slot)
        card = render_p6_card(product, product_image, card_box)
        try:
            page.paste(card, (card_box[0], card_box[1]), card)
        finally:
            card.close()
        return

    draw = ImageDraw.Draw(page)
    prepared_image = trim_product_image(product_image) if product_image is not None else None
    image_shape = classify_image_shape(prepared_image) if prepared_image is not None else "standard"
    image_box = get_product_image_box(slot, product.is_new, image_shape=image_shape)
    fitted_image = fit_product_image(prepared_image, (image_box[2] - image_box[0], image_box[3] - image_box[1]), image_shape=image_shape)
    page.paste(fitted_image, (image_box[0], image_box[1]), fitted_image)

    badge_size = get_scaled_badge_size(slot)
    badge_margin = get_scaled_badge_margin(slot)
    badge = build_por_badge(product.por_text, badge_size)
    badge_x = slot[2] - badge_size[0] - badge_margin[0]
    badge_y = slot[1] + badge_margin[1]
    page.paste(badge, (badge_x, badge_y), badge)

    if product.is_new:
        new_badge_size = get_scaled_new_badge_size(slot)
        new_badge = build_new_badge(new_badge_size)
        if new_badge is not None:
            new_badge_x = slot[0] + badge_margin[0]
            new_badge_y = slot[1] + scale_length(REFERENCE_NEW_BADGE_MARGIN_TOP, slot, "y", minimum=8)
            page.paste(new_badge, (new_badge_x, new_badge_y), new_badge)

    title_box = get_title_box(slot)
    title_font, title_lines, line_height = fit_title(draw, product.title, title_box)
    total_height = line_height * len(title_lines)
    title_raise = max(0, round((title_box[3] - title_box[1]) * 0.15))
    text_y = max(title_box[1], title_box[3] - total_height - title_raise)
    text_x = (title_box[0] + title_box[2]) // 2

    for line in title_lines:
        draw.text((text_x, text_y), line, anchor="ma", fill=TITLE_COLOR, font=title_font)
        text_y += line_height


def compose_page(
    base_page: Image.Image,
    slots: list[tuple[int, int, int, int]],
    products: list[Product],
    excel_path: Path,
    template_variant: TemplateVariant = "default",
    status_callback: Callable[[str], None] | None = None,
    progress_step: Callable[[], None] | None = None,
) -> Image.Image:
    page = base_page.copy()
    total_products = len(products)
    for index, (slot, product) in enumerate(zip(slots, products), start=1):
        if status_callback and (index == 1 or index == total_products or index % 25 == 0):
            status_callback(f"Loading image for row {product.excel_row}...")
        image = load_product_image(excel_path, product.image_source)
        draw_product(page, product, image, slot, template_variant=template_variant)
        if image is not None:
            image.close()
        if progress_step:
            progress_step()
    return page


def close_unique_images(*images: Image.Image) -> None:
    seen_ids: set[int] = set()
    for image in images:
        if id(image) in seen_ids:
            continue
        seen_ids.add(id(image))
        image.close()


def to_pdf_rect(box: tuple[int, int, int, int], page_height: int) -> tuple[int, int, int, int]:
    left, top, right, bottom = box
    width = right - left
    height = bottom - top
    return left, page_height - bottom, width, height


def scale_page_size(page_size: tuple[int, int], factor: int) -> tuple[int, int]:
    return page_size[0] * factor, page_size[1] * factor


def get_page_slots_for_role(role: str, page_size: tuple[int, int]) -> list[tuple[int, int, int, int]]:
    if role == "first":
        return get_first_page_slots(page_size)
    if role == "middle":
        return get_middle_page_slots(page_size)
    return get_last_page_slots(page_size)


def draw_cleanup_overlay_pdf(
    page_canvas: canvas.Canvas,
    page_size: tuple[int, int],
    slots: list[tuple[int, int, int, int]],
) -> None:
    inset_x = max(10, round(page_size[0] * 0.005))
    inset_y = max(10, round(page_size[1] * 0.005))
    page_canvas.setFillColor(HexColor("#FFFFFF"))
    page_canvas.setStrokeColor(HexColor("#FFFFFF"))
    for left, top, right, bottom in slots:
        x, y, width, height = to_pdf_rect(
            (left + inset_x, top + inset_y, right - inset_x, bottom - inset_y),
            page_size[1],
        )
        page_canvas.rect(x, y, width, height, stroke=0, fill=1)


def wrap_text_pdf(text: str, font_name: str, font_size: int, max_width: int) -> list[str]:
    words = text.split()
    if not words:
        return [""]
    lines: list[str] = []
    current = words[0]
    for word in words[1:]:
        trial = f"{current} {word}"
        if pdfmetrics.stringWidth(trial, font_name, font_size) <= max_width:
            current = trial
        else:
            lines.append(current)
            current = word
    lines.append(current)
    return lines


def fit_title_pdf(text: str, box: tuple[int, int, int, int]) -> tuple[str, int, list[str], int]:
    display_text = re.sub(r"\s+", " ", text.strip()).upper() or "UNTITLED PRODUCT"
    max_width = box[2] - box[0]
    max_height = box[3] - box[1]
    font_name = "Helvetica-Bold"
    start_size = max(20, round(max_height * 0.28))
    end_size = max(14, round(max_height * 0.12))

    for size in range(start_size, end_size - 1, -1):
        lines = wrap_text_pdf(display_text, font_name, size, max_width)
        if len(lines) > 3:
            continue
        line_height = max(4, size // 6) + size
        if line_height * len(lines) <= max_height:
            return font_name, size, lines, line_height

    lines = wrap_text_pdf(display_text, font_name, end_size, max_width)[:3]
    return font_name, end_size, lines, max(4, end_size // 6) + end_size


def fit_single_line_pdf(text: str, box: tuple[int, int, int, int], font_name: str = "Helvetica-Bold") -> int:
    max_width = max(1, box[2] - box[0])
    max_height = max(1, box[3] - box[1])
    start_size = max(18, round(max_height * 0.62))
    end_size = max(12, round(max_height * 0.28))
    for size in range(start_size, end_size - 1, -1):
        if pdfmetrics.stringWidth(text, font_name, size) <= max_width and size <= max_height:
            return size
    return end_size


def draw_image_on_canvas(
    page_canvas: canvas.Canvas,
    image: Image.Image,
    box: tuple[int, int, int, int],
    page_height: int,
    jpeg_quality: int | None = None,
) -> None:
    x, y, width, height = to_pdf_rect(box, page_height)
    image_buffer = io.BytesIO()
    if jpeg_quality is not None:
        rgb = Image.new("RGB", image.size, (255, 255, 255))
        if image.mode == "RGBA":
            rgb.paste(image, mask=image.split()[3])
        else:
            rgb.paste(image.convert("RGB"))
        rgb.save(image_buffer, format="JPEG", quality=jpeg_quality, optimize=True)
        image_buffer.seek(0)
        page_canvas.drawImage(ImageReader(image_buffer), x, y, width=width, height=height, mask=None)
    else:
        save_image = image
        if image.mode not in ("RGB", "RGBA", "L"):
            save_image = image.convert("RGBA")
        save_image.save(image_buffer, format="PNG")
        image_buffer.seek(0)
        page_canvas.drawImage(ImageReader(image_buffer), x, y, width=width, height=height, mask="auto")


def draw_product_pdf(
    page_canvas: canvas.Canvas,
    product: Product,
    product_image: Image.Image | None,
    slot: tuple[int, int, int, int],
    page_size: tuple[int, int],
    template_variant: TemplateVariant = "default",
) -> None:
    if template_variant == "p6":
        card_box = get_p6_card_box(slot)
        card = render_p6_card(product, product_image, card_box)
        try:
            draw_image_on_canvas(page_canvas, card, card_box, page_size[1], jpeg_quality=80)
        finally:
            card.close()
        return

    prepared_image = trim_product_image(product_image) if product_image is not None else None
    image_shape = classify_image_shape(prepared_image) if prepared_image is not None else "standard"
    image_box = get_product_image_box(slot, product.is_new, image_shape=image_shape)
    fitted_image = fit_product_image(
        prepared_image,
        (image_box[2] - image_box[0], image_box[3] - image_box[1]),
        image_shape=image_shape,
    )
    try:
        draw_image_on_canvas(page_canvas, fitted_image, image_box, page_size[1], jpeg_quality=80)
    finally:
        fitted_image.close()
        if prepared_image is not None:
            prepared_image.close()

    badge_size = get_scaled_badge_size(slot)
    badge_margin = get_scaled_badge_margin(slot)
    badge = build_por_badge(product.por_text, badge_size)
    try:
        badge_x = slot[2] - badge_size[0] - badge_margin[0]
        badge_y = slot[1] + badge_margin[1]
        draw_image_on_canvas(page_canvas, badge, (badge_x, badge_y, badge_x + badge_size[0], badge_y + badge_size[1]), page_size[1])
    finally:
        badge.close()

    if product.is_new:
        new_badge_size = get_scaled_new_badge_size(slot)
        new_badge = build_new_badge(new_badge_size)
        if new_badge is not None:
            try:
                new_badge_x = slot[0] + badge_margin[0]
                new_badge_y = slot[1] + scale_length(REFERENCE_NEW_BADGE_MARGIN_TOP, slot, "y", minimum=8)
                draw_image_on_canvas(
                    page_canvas,
                    new_badge,
                    (new_badge_x, new_badge_y, new_badge_x + new_badge_size[0], new_badge_y + new_badge_size[1]),
                    page_size[1],
                )
            finally:
                new_badge.close()

    title_box = get_title_box(slot)
    font_name, font_size, title_lines, line_height = fit_title_pdf(product.title, title_box)
    total_height = line_height * len(title_lines)
    title_raise = max(0, round((title_box[3] - title_box[1]) * 0.15))
    text_top = max(title_box[1], title_box[3] - total_height - title_raise)
    text_x = (title_box[0] + title_box[2]) / 2
    page_canvas.setFillColor(HexColor(TITLE_COLOR))
    page_canvas.setFont(font_name, font_size)
    for index, line in enumerate(title_lines):
        baseline_y = page_size[1] - (text_top + (index * line_height) + font_size)
        page_canvas.drawCentredString(text_x, baseline_y, line)


def build_overlay_pdf_bytes(
    products: list[Product],
    slots: list[tuple[int, int, int, int]],
    page_size: tuple[int, int],
    excel_path: Path,
    template_variant: TemplateVariant = "default",
    status_callback: Callable[[str], None] | None = None,
    progress_step: Callable[[], None] | None = None,
    border_style: tuple[tuple[int, int, int], int] | None = None,
    band_bottom_override: int | None = None,
) -> bytes:
    buffer = io.BytesIO()
    page_canvas = canvas.Canvas(buffer, pagesize=page_size)
    if border_style is not None:
        draw_band_cleanup_overlay_pdf(page_canvas, page_size, slots, bottom_override=band_bottom_override)
        draw_grid_borders_pdf(page_canvas, slots, border_style[0], border_style[1], page_size[1])
    else:
        draw_cleanup_overlay_pdf(page_canvas, page_size, slots)

    total_products = len(products)
    for index, (slot, product) in enumerate(zip(slots, products), start=1):
        if status_callback and (index == 1 or index == total_products or index % 25 == 0):
            status_callback(f"Loading image for row {product.excel_row}...")
        image = load_product_image(excel_path, product.image_source)
        try:
            draw_product_pdf(page_canvas, product, image, slot, page_size, template_variant=template_variant)
        finally:
            if image is not None:
                image.close()
        if progress_step:
            progress_step()

    page_canvas.showPage()
    page_canvas.save()
    return buffer.getvalue()


def split_remaining_products_for_last_page(
    products: list[Product],
    middle_capacity: int,
    last_capacity: int,
) -> tuple[list[list[Product]], list[Product], Literal["footer_safe", "last"]]:
    if not products:
        return [], [], "last"

    safe_middle_capacity = max(1, middle_capacity)
    safe_last_capacity = max(1, last_capacity)
    remainder = len(products) % safe_middle_capacity

    # Keep all non-final pages full. If the smaller last-page grid cannot absorb the
    # remainder cleanly, expand the last page by one row (footer-safe layout) instead.
    if remainder == 0 or remainder > safe_last_capacity:
        final_layout = "footer_safe"
        last_page_count = remainder or min(len(products), safe_middle_capacity)
    else:
        final_layout = "last"
        last_page_count = remainder

    if len(products) <= last_page_count:
        return [], products, final_layout

    middle_products = products[:-last_page_count]
    last_products = products[-last_page_count:]
    middle_chunks = [
        middle_products[offset : offset + safe_middle_capacity]
        for offset in range(0, len(middle_products), safe_middle_capacity)
    ]
    return middle_chunks, last_products, final_layout


def split_remaining_products_for_reserved_last_page(
    products: list[Product],
    middle_capacity: int,
    last_capacity: int,
) -> tuple[list[list[Product]], list[Product]]:
    if not products:
        return [], []

    safe_middle_capacity = max(1, middle_capacity)
    safe_last_capacity = max(1, last_capacity)
    last_page_count = min(len(products), safe_last_capacity)
    middle_products = products[:-last_page_count]
    last_products = products[-last_page_count:]
    middle_chunks = [
        middle_products[offset : offset + safe_middle_capacity]
        for offset in range(0, len(middle_products), safe_middle_capacity)
    ]
    return middle_chunks, last_products


def export_catalog_high_quality(
    excel_path: Path,
    template_pdf: Path,
    output_path: Path | None = None,
    status_callback: Callable[[str], None] | None = None,
    progress_callback: Callable[[int], None] | None = None,
    template_page_numbers: tuple[int, int, int] | None = None,
    template_form: TemplateFormName = "auto",
) -> Path:
    template_variant = resolve_template_variant(excel_path, template_pdf, template_form=template_form)
    last_progress_value = -1

    def update_progress(value: int) -> None:
        nonlocal last_progress_value
        clamped = max(0, min(100, value))
        if clamped == last_progress_value:
            return
        last_progress_value = clamped
        if progress_callback:
            progress_callback(clamped)

    if status_callback:
        status_callback("Reading Excel file...")
    update_progress(0)
    products = load_products(excel_path, template_variant=template_variant)
    total_units = max(1, len(products) + 3)
    completed_units = 1
    update_progress(round((completed_units / total_units) * 100))

    first_index, middle_index, last_index = template_page_numbers or resolve_template_page_numbers(
        template_pdf,
        status_callback=status_callback,
    )
    first_raw = render_pdf_page(template_pdf, first_index)
    middle_raw = render_pdf_page(template_pdf, middle_index)
    last_raw = render_pdf_page(template_pdf, last_index)
    completed_units += 1
    update_progress(round((completed_units / total_units) * 100))

    first_page_size = scale_page_size(first_raw.size, HIGH_QUALITY_OVERLAY_SCALE)
    middle_page_size = scale_page_size(middle_raw.size, HIGH_QUALITY_OVERLAY_SCALE)
    last_page_size = scale_page_size(last_raw.size, HIGH_QUALITY_OVERLAY_SCALE)
    first_slots = get_template_slots(template_variant, "first", first_page_size)
    middle_slots = get_template_slots(template_variant, "middle", middle_page_size)
    last_slots = get_template_slots(template_variant, "last", last_page_size)
    footer_top_reference = detect_footer_top_reference(last_raw)
    footer_safe_border_style = detect_grid_border_style(
        last_raw, get_template_slots(template_variant, "last", last_raw.size)
    ) or detect_grid_border_style(middle_raw, get_template_slots(template_variant, "middle", middle_raw.size))
    if footer_safe_border_style is not None:
        border_color, border_thickness = footer_safe_border_style
        footer_safe_border_style = (border_color, max(2, round(border_thickness * HIGH_QUALITY_OVERLAY_SCALE)))
    close_unique_images(first_raw, middle_raw, last_raw)

    final_output = build_output_path(excel_path, template_pdf, output_path)
    final_output.parent.mkdir(parents=True, exist_ok=True)
    cursor = 0
    writer = PdfWriter()
    reader = PdfReader(str(template_pdf))

    def advance_product_progress() -> None:
        nonlocal completed_units
        completed_units += 1
        update_progress(round((completed_units / total_units) * 100))

    def append_page(
        label: str,
        source_page_number: int,
        page_size: tuple[int, int],
        slots: list[tuple[int, int, int, int]],
        chunk: list[Product],
        border_style: tuple[tuple[int, int, int], int] | None = None,
        band_bottom_override: int | None = None,
    ) -> None:
        if status_callback:
            status_callback(label)
        overlay_pdf = build_overlay_pdf_bytes(
            chunk,
            slots,
            page_size,
            excel_path,
            template_variant=template_variant,
            status_callback=status_callback,
            progress_step=advance_product_progress,
            border_style=border_style,
            band_bottom_override=band_bottom_override,
        )
        writer.add_page(reader.pages[source_page_number - 1])
        page = writer.pages[-1]
        overlay_reader = PdfReader(io.BytesIO(overlay_pdf))
        overlay_page = overlay_reader.pages[0]
        width = float(page.mediabox.width)
        height = float(page.mediabox.height)
        transform = Transformation().scale(width / page_size[0], height / page_size[1])
        page.merge_transformed_page(overlay_page, transform)

    first_chunk = products[cursor : cursor + len(first_slots)]
    append_page("Composing first page...", first_index, first_page_size, first_slots, first_chunk)
    cursor += len(first_slots)

    remaining_products = products[cursor:]
    if remaining_products:
        final_layout: Literal["footer_safe", "last"] = "last"
        if template_variant == "p6":
            middle_chunks, last_products = split_remaining_products_for_reserved_last_page(
                remaining_products,
                len(middle_slots),
                len(last_slots),
            )
            final_slots = last_slots
        else:
            middle_chunks, last_products, final_layout = split_remaining_products_for_last_page(
                remaining_products,
                len(middle_slots),
                len(last_slots),
            )
            final_slots = (
                last_slots
                if final_layout == "last"
                else get_footer_safe_last_page_slots(last_page_size, footer_top_reference)
            )

        for page_index, chunk in enumerate(middle_chunks, start=1):
            append_page(
                f"Composing middle page {page_index}...",
                middle_index,
                middle_page_size,
                middle_slots,
                chunk,
            )

        last_border_style = footer_safe_border_style if final_layout == "footer_safe" else None
        last_band_bottom_override = (
            max(slot[3] for slot in last_slots) if final_layout == "footer_safe" else None
        )
        append_page(
            "Composing last page...",
            last_index,
            last_page_size,
            final_slots,
            last_products,
            last_border_style,
            last_band_bottom_override,
        )

    if status_callback:
        status_callback("Saving high-quality PDF...")
    writer.compress_identical_objects(remove_identicals=True, remove_orphans=True)
    with final_output.open("wb") as handle:
        writer.write(handle)
    update_progress(100)
    if status_callback:
        status_callback("Export complete.")
    return final_output.resolve()


def export_catalog_normal(
    excel_path: Path,
    template_pdf: Path,
    output_path: Path | None = None,
    status_callback: Callable[[str], None] | None = None,
    progress_callback: Callable[[int], None] | None = None,
    template_page_numbers: tuple[int, int, int] | None = None,
    template_form: TemplateFormName = "auto",
) -> Path:
    template_variant = resolve_template_variant(excel_path, template_pdf, template_form=template_form)
    last_progress_value = -1

    def update_progress(value: int) -> None:
        nonlocal last_progress_value
        clamped = max(0, min(100, value))
        if clamped == last_progress_value:
            return
        last_progress_value = clamped
        if progress_callback:
            progress_callback(clamped)

    if status_callback:
        status_callback("Reading Excel file...")
    update_progress(0)
    products = load_products(excel_path, template_variant=template_variant)
    total_units = max(1, len(products) + 3)
    completed_units = 1
    update_progress(round((completed_units / total_units) * 100))
    first_raw, middle_raw, last_raw = load_template_themes(
        template_pdf,
        status_callback=status_callback,
        page_numbers=template_page_numbers,
    )
    completed_units += 1
    update_progress(round((completed_units / total_units) * 100))

    first_slots = get_template_slots(template_variant, "first", first_raw.size)
    middle_slots = get_template_slots(template_variant, "middle", middle_raw.size)
    last_slots = get_template_slots(template_variant, "last", last_raw.size)
    footer_top_reference = detect_footer_top_reference(last_raw)

    first_theme = sanitize_template_page(first_raw, first_slots)
    middle_theme = sanitize_template_page(middle_raw, middle_slots)
    remaining_products = products[len(first_slots) :]
    final_layout: Literal["footer_safe", "last"] = "last"
    if template_variant == "p6":
        final_slots = last_slots
    else:
        _, _, final_layout = split_remaining_products_for_last_page(
            remaining_products,
            len(middle_slots),
            len(last_slots),
        )
        final_slots = (
            last_slots
            if final_layout == "last"
            else get_footer_safe_last_page_slots(last_raw.size, footer_top_reference)
        )
    if final_layout == "footer_safe":
        border_style = detect_grid_border_style(last_raw, last_slots) or detect_grid_border_style(
            middle_raw, middle_slots
        )
        band_bottom_override = max(slot[3] for slot in last_slots)
        last_theme = clear_grid_band(last_raw, final_slots, bottom_override=band_bottom_override)
        if border_style is not None:
            draw_grid_borders(last_theme, final_slots, *border_style)
    else:
        last_theme = sanitize_template_page(last_raw, final_slots)
    close_unique_images(first_raw, middle_raw, last_raw)

    final_output = build_output_path(excel_path, template_pdf, output_path)
    final_output.parent.mkdir(parents=True, exist_ok=True)
    cursor = 0
    page_pdf_dir = PDF_CACHE_DIR / f"rendered_{int(time.time() * 1000)}"
    clear_directory(page_pdf_dir)
    rendered_page_pdfs: list[Path] = []

    def advance_product_progress() -> None:
        nonlocal completed_units
        completed_units += 1
        update_progress(round((completed_units / total_units) * 100))

    def render_and_store_page(
        label: str,
        base_page: Image.Image,
        slots: list[tuple[int, int, int, int]],
        chunk: list[Product],
    ) -> None:
        if status_callback:
            status_callback(label)
        page_image = compose_page(
            base_page,
            slots,
            chunk,
            excel_path,
            template_variant=template_variant,
            status_callback=status_callback,
            progress_step=advance_product_progress,
        )
        page_index = len(rendered_page_pdfs) + 1
        page_pdf_path = page_pdf_dir / f"page_{page_index:04d}.pdf"
        try:
            save_page_pdf(page_image, page_pdf_path)
            rendered_page_pdfs.append(page_pdf_path)
        finally:
            page_image.close()

    try:
        first_chunk = products[cursor : cursor + len(first_slots)]
        render_and_store_page("Composing first page...", first_theme, first_slots, first_chunk)
        cursor += len(first_slots)

        remaining_products = products[cursor:]
        if remaining_products:
            if template_variant == "p6":
                middle_chunks, last_products = split_remaining_products_for_reserved_last_page(
                    remaining_products,
                    len(middle_slots),
                    len(last_slots),
                )
            else:
                middle_chunks, last_products, _final_layout = split_remaining_products_for_last_page(
                    remaining_products,
                    len(middle_slots),
                    len(last_slots),
                )

            for page_index, chunk in enumerate(middle_chunks, start=1):
                render_and_store_page(
                    f"Composing middle page {page_index}...",
                    middle_theme,
                    middle_slots,
                    chunk,
                )

            render_and_store_page("Composing last page...", last_theme, final_slots, last_products)

        if status_callback:
            status_callback("Saving PDF...")
        merge_pdf_files(rendered_page_pdfs, final_output)
        completed_units = total_units
        update_progress(100)

        if status_callback:
            status_callback("Export complete.")
        return final_output.resolve()
    finally:
        close_unique_images(first_theme, middle_theme, last_theme)
        shutil.rmtree(page_pdf_dir, ignore_errors=True)


def export_catalog(
    excel_path: Path,
    template_pdf: Path,
    output_path: Path | None = None,
    status_callback: Callable[[str], None] | None = None,
    progress_callback: Callable[[int], None] | None = None,
    quality: ExportQuality = "normal",
    template_page_numbers: tuple[int, int, int] | None = None,
    template_form: TemplateFormName = "auto",
) -> Path:
    if quality == "high":
        return export_catalog_high_quality(
            excel_path,
            template_pdf,
            output_path=output_path,
            status_callback=status_callback,
            progress_callback=progress_callback,
            template_page_numbers=template_page_numbers,
            template_form=template_form,
        )
    return export_catalog_normal(
        excel_path,
        template_pdf,
        output_path=output_path,
        status_callback=status_callback,
        progress_callback=progress_callback,
        template_page_numbers=template_page_numbers,
        template_form=template_form,
    )
